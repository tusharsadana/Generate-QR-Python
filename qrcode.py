# -*- coding: utf-8 -*-
"""
Created on Tue Oct 10 11:04:21 2017

@author: Tushar
"""

import openpyxl
import pyqrcode





wb2 = openpyxl.load_workbook('team.xlsx')
ws = wb2.active



for row in range(2,14):
    name = str( ws[('A'+str(row))].value)
    mob =  str(ws[('E'+str(row))].value)
    email = str( ws[('F'+str(row))].value)
    reg = str( ws[('B'+str(row))].value)
    des = str( ws[('G'+str(row))].value)
    qr = pyqrcode.create( name +"," +" " + reg +"," +" " + mob +"," +" " + email+"," +" " +des)
    qr.png(reg+'.png',scale = 7)



